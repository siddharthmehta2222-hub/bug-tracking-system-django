from urllib import request

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponseForbidden, HttpResponse
from django.db.models import Q
from django.contrib.auth.hashers import make_password
from .models import Notification, Project, Bug
from django.core.paginator import Paginator
from .models import Bug
from .forms import BugForm, UserSignupForm
import openpyxl
from reportlab.pdfgen import canvas


User = get_user_model()


# ==================================================
# ADMIN ONLY DECORATOR
# ==================================================
def admin_only(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.role != 'admin':
            return HttpResponseForbidden("You are not allowed ❌")
        return view_func(request, *args, **kwargs)
    return wrapper


# ==================================================
# USER SIGNUP
# ==================================================
def userSignupView(request):
    print("METHOD:", request.method)   # 🔥 DEBUG

    if request.method == "POST":
        print("POST DATA:", request.POST)   # 🔥 DEBUG

        form = UserSignupForm(request.POST)

        if form.is_valid():
            print("FORM VALID ✅")   # 🔥 DEBUG
            form.save()
            messages.success(request, "Account created successfully. Please login.")
            return redirect('login')
        else:
            print("FORM ERRORS ❌:", form.errors)   # 🔥 DEBUG

    else:
        form = UserSignupForm()

    return render(request, 'core/signup.html', {'form': form})
# ==================================================
# LOGIN
# ==================================================
def loginView(request):
    next_url = request.GET.get('next')
    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")

        try:
            user_obj = User.objects.get(email=email)
            user = authenticate(
                request,
                username=user_obj.username,   # 🔥 FIX HERE
                password=password
            )
        except User.DoesNotExist:
            user = None

        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, "Invalid email or password ❌")

    return render(request, "core/login.html")
# LOGOUT
# ==================================================
@login_required
def logoutView(request):
    logout(request)
    return redirect('login')


# ==================================================
# 🚀 DASHBOARD (FIXED + UPGRADED)
# ==================================================
@login_required
def dashboardView(request):

    user = request.user

    # ==================================================
    # 🔔 NEW: NOTIFICATION COUNT (ADDED)
    # ==================================================
    from .models import Notification
    unread_count = Notification.objects.filter(
        user=user,
        is_read=False
    ).count()


    # ==================================================
    # 🎯 ROLE FLAGS (🔥 NEW - IMPORTANT FOR TEMPLATE)
    # ==================================================
    is_admin = user.role == 'admin'
    is_manager = user.role == 'manager'
    is_developer = user.role == 'developer'
    is_tester = user.role == 'tester'


    # ==================================================
    # 🔥 ROLE BASED DATA (IMPROVED)
    # ==================================================
    if is_admin or is_manager:
        bugs = Bug.objects.all()

    elif is_developer:
        bugs = Bug.objects.filter(assigned_to=user)

    elif is_tester:
        bugs = Bug.objects.filter(reported_by=user)

    else:
        bugs = Bug.objects.none()


    # ==================================================
    # 🔍 SEARCH + FILTER (SAFE DEFAULT)
    # ==================================================
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')

    if search_query:
        bugs = bugs.filter(title__icontains=search_query)

    if status_filter:
        bugs = bugs.filter(status=status_filter)


    # ==================================================
    # 📊 BUG COUNTS (ROLE SAFE)
    # ==================================================
    total_bugs = bugs.count()
    open_bugs = bugs.filter(status='open').count()
    progress_bugs = bugs.filter(status='in_progress').count()
    closed_bugs = bugs.filter(status='closed').count()


    # ==================================================
    # 📁 PROJECT COUNTS
    # ==================================================
    total_projects = Project.objects.count()
    active_projects = Project.objects.filter(status='active').count()
    completed_projects = Project.objects.filter(status='completed').count()
    on_hold_projects = Project.objects.filter(status='on_hold').count()


    # ==================================================
    # 👥 USER COUNT (ADMIN / MANAGER ONLY)
    # ==================================================
    from django.contrib.auth import get_user_model
    User = get_user_model()

    total_users = User.objects.count() if (is_admin or is_manager) else 0


    # ==================================================
    # 🔥 LATEST BUGS (SMART BASED ON ROLE)
    # ==================================================
    latest_bugs = bugs.order_by('-id')[:5]


    # ==================================================
    # 🎯 ROLE BASED EXTRA DATA
    # ==================================================
    developer_bugs = Bug.objects.filter(assigned_to=user).count() if is_developer else 0
    tester_bugs = Bug.objects.filter(reported_by=user).count() if is_tester else 0


    # ==================================================
    # 🚀 NEW: PRIORITY COUNTS (ADVANCED UI)
    # ==================================================
    high_priority = bugs.filter(priority='high').count()
    medium_priority = bugs.filter(priority='medium').count()
    low_priority = bugs.filter(priority='low').count()
    critical_priority = bugs.filter(priority='critical').count()


    # ==================================================
    # 🚀 NEW: STATUS % (FOR PROGRESS BARS / FUTURE UI)
    # ==================================================
    total = total_bugs if total_bugs > 0 else 1

    open_percent = int((open_bugs / total) * 100)
    progress_percent = int((progress_bugs / total) * 100)
    closed_percent = int((closed_bugs / total) * 100)


    # ==================================================
    # 🎯 FINAL CONTEXT (UPGRADED)
    # ==================================================
    context = {
        # ROLE
        'role': user.role,
        'is_admin': is_admin,
        'is_manager': is_manager,
        'is_developer': is_developer,
        'is_tester': is_tester,

        # BUG DATA
        'total_bugs': total_bugs,
        'open_bugs': open_bugs,
        'progress_bugs': progress_bugs,
        'closed_bugs': closed_bugs,
        'bugs': latest_bugs,

        # PROJECT DATA
        'total_projects': total_projects,
        'active_projects': active_projects,
        'completed_projects': completed_projects,
        'on_hold_projects': on_hold_projects,

        # USER DATA
        'total_users': total_users,

        # ROLE SPECIFIC
        'developer_bugs': developer_bugs,
        'tester_bugs': tester_bugs,

        # SEARCH
        'search_query': search_query,
        'status_filter': status_filter,

        # 🚀 ADVANCED DATA
        'high_priority': high_priority,
        'medium_priority': medium_priority,
        'low_priority': low_priority,
        'critical_priority': critical_priority,

        'open_percent': open_percent,
        'progress_percent': progress_percent,
        'closed_percent': closed_percent,

        # 🔔 NEW (IMPORTANT)
        'unread_count': unread_count,
    }

    return render(request, "core/dashboard.html", context)
# ==================================================
# 🔥 FIXED STATUS FUNCTION
# ==================================================
@login_required
def change_status(request, id, status):
    bug = get_object_or_404(Bug, id=id)
    bug.status = status
    bug.save()
    return redirect('dashboard')


# ==================================================
# 🔥 FIXED DELETE BUG FUNCTION
# ==================================================
@login_required
def delete_bug(request, id):
    bug = get_object_or_404(Bug, id=id)

    # only admin & manager allowed
    if request.user.role in ['admin', 'manager']:
        bug.delete()

    return redirect('dashboard')
# ==================================================
# BUG ACTIONS
# ==================================================
@login_required
def start_bug(request, id):
    bug = get_object_or_404(Bug, id=id)

    # 🔐 SECURITY CHECK
    if request.user.role not in ['developer', 'manager', 'admin']:
        return HttpResponseForbidden("You are not allowed ❌")

    # ✅ Only allow if assigned OR admin/manager
    if request.user.role == 'developer' and bug.assigned_to != request.user:
        return HttpResponseForbidden("Not your bug ❌")

    if bug.status == 'open':
        bug.status = "in_progress"
        bug.save()

    messages.success(request, "Bug moved to In Progress ✅")
    return redirect('dashboard')

@login_required
def close_bug(request, id):
    bug = get_object_or_404(Bug, id=id)

    # 🔐 SECURITY CHECK
    if request.user.role not in ['manager', 'admin']:
        return HttpResponseForbidden("You are not allowed ❌")

    if bug.status != 'closed':
        bug.status = "closed"
        bug.save()

    messages.success(request, "Bug closed successfully ✅")
    return redirect('dashboard')

# ==================================================
# ADD BUG
# ==================================================

@login_required
def add_bug(request):
    from .models import Project
    projects = Project.objects.all()   # ✅ GET ALL PROJECTS

    if request.method == "POST":
        title = request.POST.get('bug_name') or "Untitled Bug"

        # ✅ FIX PROJECT (IMPORTANT)
        from .models import Project
        project_id = request.POST.get('project')

        if project_id:
            try:
                project = Project.objects.get(id=project_id)
            except Project.DoesNotExist:
                project = Project.objects.first()
        else:
            project = Project.objects.first()

        tester_code = request.POST.get('tester_code')
        bug_date = request.POST.get('bug_date')
        bug_level = request.POST.get('bug_level')
        bug_priority = request.POST.get('bug_priority') or "low"
        bug_type = request.POST.get('bug_type')
        description = request.POST.get('description')

        # 🔥 NEW: IMAGE SUPPORT
        image = request.FILES.get('image')

        # 🔥 CREATE BUG (UPGRADED)
        bug = Bug.objects.create(
            title=title,
            description=description,
            priority=bug_priority,
            status='open',
            project=project,

            # 🔥 NEW FIELDS SAVED
            bug_level=bug_level,
            bug_type=bug_type,
            bug_date=bug_date,
            image=image
        )

        # ✅ Reporter
        bug.reported_by = request.user

        # 🔥 AUTO ASSIGN DEVELOPER
        from django.contrib.auth import get_user_model
        User = get_user_model()

        developer = User.objects.filter(role='developer').first()

        if developer:
            bug.assigned_to = developer
        else:
            bug.assigned_to = request.user

        bug.save()   # ✅ FIXED INDENT

        # 🔔 notification for creator
        Notification.objects.create(
            user=request.user,
            message="Bug added successfully"
        )

        messages.success(request, "Bug Added Successfully")
        return redirect('bug_list')

    # 🔥🔥🔥 FINAL FIX (DO NOT REMOVE)
    return render(request, 'core/add_bug.html', {
        'projects': projects
    })

# =====================================
# 🔔 NOTIFICATIONS
# =====================================

@login_required
def notifications_view(request):
    from .models import Notification

    data = Notification.objects.filter(
        user=request.user
    ).order_by('-created_at')

    return render(request, 'core/notifications.html', {
        'notifications': data
    })


@login_required
def mark_read(request, id):
    from .models import Notification

    try:
        notif = Notification.objects.get(id=id, user=request.user)
        notif.is_read = True
        notif.save()
    except Notification.DoesNotExist:
        pass

    return redirect('notifications')
#=================================================
# ADD PROJECT
#=================================================
@login_required
def add_project(request):
    users = User.objects.all()

    if request.method == "POST":
        name = request.POST.get('name')
        submission_date = request.POST.get('submission_date')
        duration = request.POST.get('duration')
        client_name = request.POST.get('client_name')
        client_address = request.POST.get('client_address')
        phone = request.POST.get('phone')
        email = request.POST.get('email')
        department = request.POST.get('department')
        description = request.POST.get('description')
        project_lead_id = request.POST.get('project_lead')

        # 🔥 NEW: GET STATUS
        status = request.POST.get('status') or 'active'

        # ✅ FIX PROJECT LEAD
        project_lead = None
        if project_lead_id:
            try:
                project_lead = User.objects.get(id=project_lead_id)
            except User.DoesNotExist:
                project_lead = None

        # ✅ SAVE PROJECT
        Project.objects.create(
            name=name,
            submission_date=submission_date if submission_date else None,
            duration=duration,
            client_name=client_name,
            client_address=client_address,
            phone=phone,
            email=email,
            department=department,
            description=description,
            project_lead=project_lead,

            # 🔥 NEW FIELD SAVED
            status=status
        )

        messages.success(request, "Project Added Successfully ✅")
        return redirect('project_report')

    return render(request, 'core/add_project.html', {'users': users})
#===============================================
#project report
#===============================================
@login_required
def project_report(request):
    from django.db.models import Q

    query = request.GET.get('q')

    projects = Project.objects.all().order_by('-id')

    # 🔍 SEARCH (FIXED + IMPROVED)
    if query:
        projects = projects.filter(
            Q(name__icontains=query) |
            Q(client_name__icontains=query) |
            Q(email__icontains=query) |
            Q(department__icontains=query) |
            Q(project_lead__username__icontains=query) |   # ✅ FIXED
            Q(status__icontains=query)                     # 🔥 ADDED
        )

    # 📄 PAGINATION (IMPORTANT)
    paginator = Paginator(projects, 5)  # 5 per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'core/project_report.html', {
        'projects': page_obj,
        'page_obj': page_obj,
    })
#=================================================
# view project  
#=================================================
 
def view_project(request, id):
    project = Project.objects.get(id=id)
    return render(request, 'core/view_project.html', {'project': project})



# ==================================================
# EXPORT PROJECTS EXCEL
# ==================================================
@login_required
def export_projects_excel(request):
    projects = Project.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['ID', 'Name', 'Client', 'Email', 'Department'])

    for p in projects:
        ws.append([
            p.id,
            p.name,
            getattr(p, 'client_name', ''),
            getattr(p, 'email', ''),
            getattr(p, 'department', '')
        ])

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=projects.xlsx'
    wb.save(response)

    return response


# ==================================================
# EXPORT PROJECTS PDF
# ==================================================
@login_required
def export_projects_pdf(request):
    projects = Project.objects.all()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=projects.pdf'

    p_canvas = canvas.Canvas(response)

    y = 800
    for p in projects:
        text = f"{p.id} | {p.name} | {getattr(p, 'client_name', '')}"
        p_canvas.drawString(30, y, text)
        y -= 20

        if y < 50:
            p_canvas.showPage()
            y = 800

    p_canvas.save()
    return response



# ==================================================
# BUG LIST
# ==================================================
@login_required
def bug_list(request):
    query = request.GET.get('q')
    status = request.GET.get('status')

    bugs = Bug.objects.all().order_by('-id')

    # 🔍 SEARCH
    if query:
        bugs = bugs.filter(
            Q(title__icontains=query) |
            Q(description__icontains=query)
        )

    # 🎯 FILTER
    if status:
        bugs = bugs.filter(status=status)

    # 📄 PAGINATION (FIXED ✅)
    paginator = Paginator(bugs, 5)   # 5 per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'core/bug_list.html', {
        'bugs': page_obj,
        'page_obj': page_obj,
        'query': query,
        'status': status
    })

# ==================================================
#view bug
# ==================================================
@login_required
def view_bug(request, id):
    bug = get_object_or_404(Bug, id=id)
    return render(request, 'core/view_bug.html', {'bug': bug})

# ==================================================
# edit bug
# ==================================================
@login_required
def edit_bug(request, id):
    bug = get_object_or_404(Bug, id=id)

    from .models import Project
    projects = Project.objects.all()

    from django.contrib.auth import get_user_model
    User = get_user_model()
    developers = User.objects.filter(role='developer')

    if request.method == 'POST':
        bug.title = request.POST.get('title')
        bug.priority = request.POST.get('priority')
        bug.status = request.POST.get('status')

        # 🔥 EXISTING FIELDS (KEEPED)
        bug.project_id = request.POST.get('project')
        bug.assigned_to_id = request.POST.get('assigned_to')
        bug.bug_level = request.POST.get('bug_level')
        bug.bug_type = request.POST.get('bug_type')
        bug.bug_date = request.POST.get('bug_date')
        bug.description = request.POST.get('description')

        # 🔥🔥 NEW IMAGE SAVE (IMPORTANT)
        if request.FILES.get('image'):
            bug.image = request.FILES.get('image')

        bug.save()

        messages.success(request, "Bug updated successfully ✅")
        return redirect('bug_list')

    return render(request, 'core/edit_bug.html', {
        'bug': bug,
        'projects': projects,
        'developers': developers
    })
# ==================================================
# delete project
# ==================================================
@login_required
def delete_project(request, id):
    project = get_object_or_404(Project, id=id)
    project.delete()
    messages.success(request, "Project deleted successfully")
    return redirect('project_report')

# ==================================================
# edit project
# ==================================================
@login_required
def edit_project(request, id):
    project = get_object_or_404(Project, id=id)

    if request.method == "POST":
        project.name = request.POST.get('name')
        project.client_name = request.POST.get('client_name')
        project.duration = request.POST.get('duration')
        project.email = request.POST.get('email')
        project.department = request.POST.get('department')

        project.save()

        messages.success(request, "Project updated successfully ✅")
        return redirect('project_report')

    return render(request, 'core/edit_project.html', {
        'project': project,
        'users': User.objects.all()
    })

# ==================================================
# HOME
# ==================================================
def home(request):
    return render(request, 'core/home.html')


# ==================================================
# USER MANAGEMENT
# ==================================================
@login_required
def add_user(request):
    if request.method == "POST":
        user = User.objects.create_user(
            username=request.POST.get('username'),
            email=request.POST.get('email'),
            password=request.POST.get('password')
        )

        user.role = request.POST.get('role')
        user.first_name = request.POST.get('first_name')
        user.last_name = request.POST.get('last_name')
        user.phone = request.POST.get('phone')
        user.gender = request.POST.get('gender')
        user.dob = request.POST.get('dob')
        user.address = request.POST.get('address')

        if request.FILES.get('image'):
            user.image = request.FILES.get('image')

        user.save()

        messages.success(request, "User Added Successfully ✅")
        return redirect('user_list')

    return render(request, 'core/add_user.html')


def user_list(request):
    users = User.objects.all().order_by('id')

    # 🔍 SEARCH
    search_query = request.GET.get('search')
    role_filter = request.GET.get('role')

    if search_query:
        users = users.filter(
            username__icontains=search_query
        ) | users.filter(
            email__icontains=search_query
        )

    if role_filter:
        users = users.filter(role=role_filter)

    # 🔥 PAGINATION
    paginator = Paginator(users, 5)  # 5 users per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'core/user_list.html', {
        'page_obj': page_obj
    })

@login_required
def delete_user(request, id):
    user = get_object_or_404(User, id=id)
    user.delete()
    messages.success(request, "User deleted successfully")
    return redirect('user_list')


@login_required
def view_user(request, id):
    user = get_object_or_404(User, id=id)
    return render(request, 'core/view_user.html', {'user': user})


@login_required
def edit_user(request, id):
    user = get_object_or_404(User, id=id)

    if request.method == "POST":
        user.first_name = request.POST.get('first_name')
        user.last_name = request.POST.get('last_name')
        user.phone = request.POST.get('phone')
        user.gender = request.POST.get('gender')
        user.dob = request.POST.get('dob')
        user.address = request.POST.get('address')
        user.role = request.POST.get('role')

        if request.FILES.get('image'):
            user.image = request.FILES.get('image')

        user.save()
        messages.success(request, "User updated successfully ✅")
        return redirect('user_list')

    return render(request, 'core/edit_user.html', {'user': user})


# ==================================================
# CHANGE PASSWORD
# ==================================================
@login_required

def change_password(request, id):
    user = get_object_or_404(User, id=id)

    if request.method == "POST":
        password = request.POST.get('password')
        confirm = request.POST.get('confirm_password')

        if password != confirm:
            messages.error(request, "Passwords do not match ❌")
            return redirect('change_password', id=id)

        user.set_password(password)
        user.save()

        messages.success(request, "Password updated successfully ✅")
        return redirect('user_list')

    return render(request, 'core/change_password.html', {'user': user})

# ==================================================
# MY PROFILE    
# ==================================================
@login_required
def my_profile(request):
    return render(request, 'core/my_profile.html', {'user': request.user})

# ==================================================
# EXPORT USERS EXCEL
# ==================================================
@login_required
def export_users_excel(request):
    users = User.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['ID', 'Name', 'Email', 'Phone', 'Role'])

    for user in users:
        ws.append([
            user.id,
            f"{user.first_name} {user.last_name}",
            user.email,
            user.phone,
            user.role
        ])

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=users.xlsx'
    wb.save(response)

    return response


# ==================================================
# EXPORT USERS PDF
# ==================================================
@login_required
def export_users_pdf(request):
    users = User.objects.all()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=users.pdf'

    p = canvas.Canvas(response)

    y = 800
    for user in users:
        p.drawString(30, y, f"{user.id} | {user.email} | {user.role}")
        y -= 20

    p.save()
    return response


# ==================================================
# EXPORT BUGS EXCEL
# ==================================================
@login_required
def export_bugs_excel(request):
    bugs = Bug.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['ID', 'Title', 'Priority', 'Status'])

    for bug in bugs:
        ws.append([
            bug.id,
            bug.title,
            bug.priority,
            bug.status
        ])

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=bugs.xlsx'
    wb.save(response)

    return response


# ==================================================
# EXPORT BUGS PDF
# ==================================================
@login_required
def export_bugs_pdf(request):
    bugs = Bug.objects.all()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=bugs.pdf'

    p = canvas.Canvas(response)

    y = 800
    for bug in bugs:
        text = f"{bug.id} | {bug.title} | {bug.status}"
        p.drawString(30, y, text)
        y -= 20

        if y < 50:
            p.showPage()
            y = 800

    p.save()
    return response


# ==================================================
# USER REPORT
# ==================================================
@login_required
def user_report(request):
    users = User.objects.all().order_by('-id')
    return render(request, 'core/user_report.html', {'users': users})

# ==================================================
# ABOUT PAGE
# ==================================================
def about(request):
    return render(request, 'core/about.html')


# ==================================================
# CONTACT PAGE
# ==================================================
def contact(request):
    return render(request, 'core/contact.html')

